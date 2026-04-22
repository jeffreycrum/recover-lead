import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import openpyxl

from app.ingestion.base_scraper import BaseScraper, RawLead
from app.ingestion.factory import register_scraper
from app.ingestion.tls import scraper_client

# Hard cap on raw XLSX byte size before we hand it to openpyxl. Even with
# read_only=True, a crafted XLSX (zip-bomb) can exhaust worker memory at
# load-time. Sized for the largest observed county publications plus headroom.
MAX_XLSX_BYTES = 50 * 1024 * 1024  # 50 MiB

# Sale types that normalizer.py / downstream templates know how to render.
# Keep aligned with the comment on Lead.sale_type in app/models/lead.py.
_ALLOWED_SALE_TYPES = frozenset(
    {"tax_deed", "foreclosure", "lien", "property_tax_refund"}
)
_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


@register_scraper("XlsxScraper")
class XlsxScraper(BaseScraper):
    """Scraper for counties that provide surplus fund lists as Excel downloads.

    Handles Hillsborough-style claims tracking sheets where owner names
    and amounts are embedded in a claims narrative column.
    """

    def __init__(
        self,
        county_name: str,
        source_url: str,
        state: str = "FL",
        config: dict | None = None,
    ):
        super().__init__(county_name, state)
        self.source_url = source_url
        self.config = config or {}

    async def fetch(self) -> bytes:
        async with scraper_client(self.source_url) as client:
            response = await client.get(self.source_url)
            response.raise_for_status()
            content = response.content
            if len(content) > MAX_XLSX_BYTES:
                raise ValueError(
                    f"XLSX response exceeds {MAX_XLSX_BYTES} bytes (got {len(content)})"
                )
            return content

    def parse(self, raw_data: bytes) -> list[RawLead]:
        """Parse Excel data into leads.

        Two modes supported via config:

        - Default (claims mode): Hillsborough-style sheet where column 1
          contains a free-text "claims narrative" and we regex out the
          largest dollar amount + first claimant name.

        - simple_table_mode: plain tabular XLSX with a header row and
          one lead per data row. Used by Madison. Config:
            {
              "simple_table_mode": True,
              "columns": {
                "case_number": 0,
                "parcel_id": 2,
                "property_address": 3,
                "surplus_amount": 4,
                "owner_name": 5,
              },
              "skip_rows_containing": ["Tax Deed Surplus List"]
            }
          A row is a valid lead only if the case_number cell is non-empty
          and the surplus_amount cell is numeric-ish — which cleanly skips
          header rows, blank spacer rows, and owner/address continuation
          rows that are common in these sheets.
        """
        if len(raw_data) > MAX_XLSX_BYTES:
            raise ValueError(
                f"XLSX exceeds {MAX_XLSX_BYTES} bytes (got {len(raw_data)})"
            )
        if self.config.get("simple_table_mode"):
            return self._parse_simple_table(raw_data)

        leads = []
        wb = openpyxl.load_workbook(
            BytesIO(raw_data),
            read_only=True,
            data_only=True,
        )
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # Skip header

            cells = []
            for c in row:
                val = str(c).strip() if c else ""
                # Strip formula injection characters
                if val and val[0] in ("=", "+", "-", "@"):
                    val = "'" + val
                cells.append(val)
            if not cells or not cells[0]:
                continue

            case_number = cells[0]

            # Extract owner name and amount from claims column
            claims_text = cells[1] if len(cells) > 1 else ""
            owner_name, surplus_amount = self._extract_from_claims(claims_text)

            if not owner_name and not surplus_amount:
                # Try case number as lead with no details
                owner_name = None
                surplus_amount = Decimal("0.00")

            if surplus_amount <= 0:
                continue

            leads.append(
                RawLead(
                    case_number=case_number,
                    owner_name=owner_name,
                    surplus_amount=surplus_amount,
                    property_state=self.state,
                    sale_type="tax_deed",
                    raw_data={"row": cells},
                )
            )

        wb.close()
        return leads

    def _parse_simple_table(self, raw_data: bytes) -> list[RawLead]:
        """Parse a plain tabular XLSX with fixed column positions."""
        col_map = self.config.get("columns") or {}
        skip_rows_containing = [
            s.lower() for s in self.config.get("skip_rows_containing", [])
        ]

        case_col = col_map.get("case_number", 0)
        amt_col = col_map.get("surplus_amount", 4)
        owner_col = col_map.get("owner_name")
        addr_col = col_map.get("property_address")
        parcel_col = col_map.get("parcel_id")
        sale_date_col = col_map.get("sale_date")
        # Config is trusted but not perfect — validate before letting it
        # leak into RawLead.sale_type, which downstream templates branch on.
        raw_sale_type = (self.config.get("sale_type") or "tax_deed").strip()
        sale_type = raw_sale_type if raw_sale_type in _ALLOWED_SALE_TYPES else "tax_deed"

        leads: list[RawLead] = []
        wb = openpyxl.load_workbook(BytesIO(raw_data), read_only=True, data_only=True)
        try:
            ws = wb.active

            def cell(row: tuple, idx: int | None) -> str:
                if idx is None or idx >= len(row) or row[idx] is None:
                    return ""
                val = str(row[idx]).strip()
                if val and val[0] in ("=", "+", "-", "@"):
                    val = "'" + val
                return val

            def sale_date_cell(row: tuple, idx: int | None) -> str | None:
                # openpyxl returns a datetime for date-formatted cells; emit
                # ISO YYYY-MM-DD so downstream Lead.sale_date (Date column)
                # parses cleanly. String cells must already be ISO — normalizer.py
                # uses date.fromisoformat and silently drops anything else, so
                # we pre-validate here to avoid persisting bogus values.
                if idx is None or idx >= len(row) or row[idx] is None:
                    return None
                raw = row[idx]
                if isinstance(raw, datetime):
                    return raw.date().isoformat()
                if isinstance(raw, date):
                    return raw.isoformat()
                text = cell(row, idx)
                if not text:
                    return None
                return text if _ISO_DATE_RE.match(text) else None

            for row in ws.iter_rows(values_only=True):
                case_number = cell(row, case_col)
                if not case_number:
                    continue

                # Skip header/title rows
                lower = case_number.lower()
                if any(skip in lower for skip in skip_rows_containing):
                    continue
                if "case" in lower and "#" in lower:
                    continue

                # Amount must be numeric — skips continuation rows and headers
                amt_raw = row[amt_col] if amt_col < len(row) else None
                if isinstance(amt_raw, (int, float)):
                    surplus_amount = Decimal(str(amt_raw))
                else:
                    surplus_amount = self._parse_amount_str(str(amt_raw or ""))
                if surplus_amount <= 0:
                    continue

                leads.append(
                    RawLead(
                        case_number=case_number,
                        parcel_id=cell(row, parcel_col) or None,
                        property_address=cell(row, addr_col) or None,
                        owner_name=cell(row, owner_col) or None,
                        surplus_amount=surplus_amount,
                        property_state=self.state,
                        sale_date=sale_date_cell(row, sale_date_col),
                        sale_type=sale_type,
                        raw_data={"row": [str(c) if c is not None else "" for c in row]},
                    )
                )
        finally:
            wb.close()
        return leads

    @staticmethod
    def _parse_amount_str(amount_str: str) -> Decimal:
        """Parse a currency-like string into a Decimal, 0 on failure."""
        if not amount_str:
            return Decimal("0.00")
        cleaned = re.sub(r"[^\d.]", "", amount_str)
        if not cleaned:
            return Decimal("0.00")
        try:
            value = Decimal(cleaned)
        except Exception:
            return Decimal("0.00")
        if value >= Decimal("10000000000") or value < 0:
            return Decimal("0.00")
        return value

    def _extract_from_claims(self, claims_text: str) -> tuple[str | None, Decimal]:
        """Extract the primary claimant name and largest claimed amount."""
        if not claims_text or claims_text.lower() in ("no claims filed", "none"):
            return None, Decimal("0.00")

        # Find all dollar amounts
        amounts = re.findall(r"\$[\d,]+\.?\d*", claims_text)
        max_amount = Decimal("0.00")
        for amt_str in amounts:
            cleaned = re.sub(r"[^\d.]", "", amt_str)
            try:
                amt = Decimal(cleaned)
                if amt > max_amount:
                    max_amount = amt
            except Exception:
                continue

        # Extract first claimant name (pattern: "N. Name Surname, date, amount")
        # Look for text after a number+period at the start of claims
        name_match = re.search(r"^\d+\.\s*(.+?)(?:,\s*\d{1,2}/)", claims_text.strip())
        owner_name = name_match.group(1).strip() if name_match else None

        return owner_name, max_amount
